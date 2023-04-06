import openpyxl
from datetime import datetime

INITIAL_SHEET_NAME = 'Sheet'

# Column names
PROCEDURE_NAME = 'Procedure Name'
CLIENT_NAME = 'Client Name'
START_DATE_TIME = 'Start Time'
END_DATE_TIME = 'End Time'
DESCRIPTION = 'Description'

# Get the current year
cur_year = str(datetime.now().year)

class ExcelWorkbook:
    ''' ExcelWorkbook will be responsible for the managment of the excel file which will be used for
    storing information about the google calendar events of Silvy Spa. The excel will have sheets
    for each year and each sheet will contain all events for the specified year by date and time '''

    def __init__(self, wb_name, year=cur_year) -> None:
        self.name = wb_name
        self.cur_sheet_name = None
        # { Column name: column letter }
        self.column_names = {PROCEDURE_NAME: 'A', CLIENT_NAME: 'B', START_DATE_TIME: 'C',
                             END_DATE_TIME: 'D', DESCRIPTION: 'E'}
        
        self.procedure_col_letter = self.column_names[PROCEDURE_NAME]
        self.client_col_letter = self.column_names[CLIENT_NAME]
        self.start_col_letter = self.column_names[START_DATE_TIME]
        self.end_col_letter = self.column_names[END_DATE_TIME]
        self.desc_col_letter = self.column_names[DESCRIPTION]

        # Initialize the workbook
        self.workbook = self._get_workbook()
        # Initialize the worksheet responsive for the specified year events
        self.worksheet = self._get_worksheet(year)
        
    def _get_workbook(self):
        ''' Load the workbook which is used for storing the information from the calendar. If it is
        not existing it will be automaticly created.'''
        # Check if the file exists
        try:
            return openpyxl.load_workbook(self.name)
        except FileNotFoundError:
            return openpyxl.Workbook()

    def _get_worksheet(self, sheet_name):
        ''' Load the sheet in which the data will be stored. Each sheet will store the information
        for a specific year and from there the name of the sheet is comming.'''
        if sheet_name == self.cur_sheet_name or sheet_name in self.workbook.sheetnames:
            self.cur_sheet_name = sheet_name
            return self.workbook.get_sheet_by_name(self.cur_sheet_name)
        else:
            self.cur_sheet_name = sheet_name
            worksheet = self.workbook.create_sheet(title=self.cur_sheet_name)

            # Removing the empty sheet created in the beginning. It is always with default name Sheet
            if INITIAL_SHEET_NAME in self.workbook.sheetnames:
                del self.workbook[INITIAL_SHEET_NAME]
            
            self._setup_headers_of_columns(worksheet)
            return worksheet

    def _setup_headers_of_columns(self, worksheet):
        ''' Each column have header on which will corespond on the information added to this column.'''
        # Add the headers to the new sheet
        for header in self.column_names:
            col_letter = self.column_names[header]
            #'{col_letter}1' means the first cell in each column (like A1, B1 etc.)
            worksheet[f'{col_letter}1'] = header

    def _format_date_time(self, date_time):
        ''' Change the format of the datetime (start_date and end_date) which is comming from Google API from
        ISO formated(%Y-%m-%dT%H:%M:%S.%fZ) to %Y-%m-%d %H:%M'''
        dt = datetime.fromisoformat(date_time)
        return dt.strftime("%Y-%m-%d %H:%M")
    
    def _get_insertion_row_for_event(self, event):
        ''' Find on which row the event have to be inserted. If the events are the most recent they 
        will be added to the last row. Otherwise they will be added on the line on which they will 
        be ordered by date.'''
        event_date = self._format_date_time(event['start'].get('dateTime', event['start'].get('date')))

        # If the event's year is different than the sheet year that means we have to create new sheet.
        # Following %Y-%m-%d %H:%M the first 4 characters are the year
        if event_date[:4] != self.cur_sheet_name:
            self.worksheet = self._get_worksheet(event_date[:4])

        last_input_row = self.worksheet.max_row
        # On the first line we have headers so we have to add 1
        first_input_row = self.worksheet.min_row + 1 

        # Check if there are no events in the sheet
        if self.worksheet.max_row == self.worksheet.min_row:
            return first_input_row
        
        # Check if the event is with the same timestamp as the last one
        last_input_date = self.worksheet[f'{self.start_col_letter}{last_input_row}'].value

        # Find the line number of the first event with this timestamp
        while event_date == last_input_date:
            if last_input_date < event_date:
                return last_input_row + 1
            else:
                last_input_row -= 1
                last_input_date = self.worksheet[f'{self.start_col_letter}{last_input_row}'].value

        # Check if the event should be inserted after the last input
        if event_date > last_input_date:
            return last_input_row + 1
        
        # Check if the event should be inserted at the beginning of the worksheet
        first_input_date = self.worksheet[f'{self.start_col_letter}{first_input_row}'].value
        if event_date <= first_input_date:
            return first_input_row
        
        # Search for the insertion row between the first and last inputs
        for row in range(first_input_row + 1, last_input_row + 1):
            current_date = self.worksheet[f'{self.start_col_letter}{row}'].value
            if current_date >= event_date:
                return row
            
        # If we haven't found an insertion row yet, insert at the end
        print(" I didn't found where to add the input so I am adding it at the end")
        return last_input_row + 1
    
    def _split_summary(self, summary):
        ''' In the summary of the calendar events there is a pattern on which we ca separate the 
        procedure which is executed and the name of the client <procedure> - <name_of_client>.
        Our function will split the summary in tuple of procedure, client.'''
        elem = summary.split('-')
        
        # In a normal situation the number of elements should be 2 - procedure name and client name
        if len(elem) == 2:
            return (elem[0].strip(), elem[1].strip())
        else:
            return (summary, None)
        
    def _is_event_exist_on_row(self, row, event):
        ''' Check is the event equal to the data written on a specific row in current worksheet.'''
        procedure_name, client_name = self._split_summary(event['summary'])

        return self.worksheet[f'{self.procedure_col_letter}{row}'].value == procedure_name and \
                self.worksheet[f'{self.client_col_letter}{row}'].value == client_name

    def save(self):
        ''' Normal Save method which will save the current status of the workbook to the file.'''
        self.workbook.save(self.name)

    def add_event(self, event):
        ''' Add event to the right place where it should be in order to be sorted by date. The event 
        have to be with the same structure as the one returned from the Google Calendar API 
        (with four columns with the specified names)'''
        row_to_insert = self._get_insertion_row_for_event(event)

        # If we are adding the event at the end it is not needed to insert a row
        if not (row_to_insert == self.worksheet.max_row + 1):
            if not self.is_event_exist(event):
                self.worksheet.insert_rows(row_to_insert)
            else:
                row_to_insert = self.get_event_row(event)
                print("Are you sure you want to overwrite the data") #TODO: Add pop-up question

        # Save the event
        procedure_name, client_name = self._split_summary(event['summary'])
        if client_name is None:
            print('The summary is not using the pattern!')

        self.worksheet[f'{self.procedure_col_letter}{row_to_insert}'] = procedure_name
        self.worksheet[f'{self.client_col_letter}{row_to_insert}'] = client_name
        self.worksheet[f'{self.start_col_letter}{row_to_insert}'] = self._format_date_time(event['start'].get('dateTime', event['start'].get('date')))
        self.worksheet[f'{self.end_col_letter}{row_to_insert}'] = self._format_date_time(event['end'].get('dateTime', event['end'].get('date')))
        self.worksheet[f'{self.desc_col_letter}{row_to_insert}'] = event.get('description', None)

    def get_last_updated_date(self):
        ''' Get the date of the last input in the specified worksheet.'''
        last_input_row = self.worksheet.max_row
        return self.worksheet[f'{self.start_col_letter}{last_input_row}'].value
    
    def is_event_exist(self, event):
        ''' Lets assume that if the summary and the start date are equal then the event exists. 
        This function will return True or False if the event exists.'''
        row_to_insert = self._get_insertion_row_for_event(event)

        event_date_time = self._format_date_time(event['start'].get('dateTime', event['start'].get('date')))
        row_date_time = self.worksheet[f'{self.start_col_letter}{row_to_insert}'].value

        # Search for the event through all events in this day and hour
        while event_date_time == row_date_time and row_to_insert != self.worksheet.max_row:
            if self._is_event_exist_on_row(row_to_insert, event):
                return True
            else:
                row_to_insert += 1
                row_date_time = self.worksheet[f'{self.start_col_letter}{row_to_insert}'].value
        
        return self._is_event_exist_on_row(row_to_insert, event)

    def get_event_row(self, event):
        ''' Lets assume that if the summary and the start date are equal then the event exists. 
        This function will return the line number of the event in the worksheet'''
        row_to_insert = self._get_insertion_row_for_event(event)

        event_date_time = self._format_date_time(event['start'].get('dateTime', event['start'].get('date')))
        row_date_time = self.worksheet[f'{self.start_col_letter}{row_to_insert}'].value

        # Search for the event through all events in this day and hour
        while event_date_time == row_date_time:
            if self._is_event_exist_on_row(row_to_insert, event):
                return row_to_insert
            else:
                row_to_insert += 1
                event_date_time = self._format_date_time(event['start'].get('dateTime', event['start'].get('date')))
                row_date_time = self.worksheet[f'{self.start_col_letter}{row_to_insert}'].value
        
        raise ValueError("Event not found")